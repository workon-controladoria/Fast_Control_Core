import os
import pandas as pd
from colorama import Fore, Style
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


from ..config import settings, account_mapping
from ..config.settings import CAMINHO_BASE_PADRAO, OUTPUT_DIR

CAMINHO_ARQUIVO_ENTRADA = CAMINHO_BASE_PADRAO
CAMINHO_PASTA_SAIDA = OUTPUT_DIR


def _determinar_conta_contabil(row):
    try:
        natureza = str(row['NATUREZA']).strip()
        classificacao = str(row['CUSTO OU DESPESA']).strip().upper()
        if 'CUSTO' in classificacao:
            return account_mapping.mapeamento_custo.get(natureza, 'Custo Não Mapeado')
        elif 'DESPESA' in classificacao:
            return account_mapping.mapeamento_despesa.get(natureza, 'Despesa Não Mapeada')
        else:
            return 'Classificação Inválida'
    except KeyError:
        return "Coluna Essencial Faltando"

def _aplicar_estilo_transitorias(ws):
    """
    Aplica a formatação visual (cores) às linhas da aba de transitorias.
    Espera que a worksheet contenha a coluna 'STATUS_CONCILIACAO' e opcionalmente 'DIFERENCA'.
    """
    # Descobrir em qual coluna caiu o "STATUS_CONCILIACAO"
    coluna_status_idx = None
    coluna_diferenca_idx = None
    
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == 'STATUS_CONCILIACAO':
            coluna_status_idx = col_idx
        elif cell.value == 'DIFERENCA':
            coluna_diferenca_idx = col_idx
            
    # Aplicar as cores se as colunas de conciliação existirem
    if coluna_status_idx:
        # Padrões de cores (Fundo e Letra)
        fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
        font_ok = Font(color="006100") 
        
        fill_div = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Vermelho
        font_div = Font(color="9C0006") 
        
        fill_nao_enc = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Amarelo
        font_nao_enc = Font(color="9C5700")
        
        # Varre as linhas pintando o status
        for row in range(2, ws.max_row + 1):
            cell_status = ws.cell(row=row, column=coluna_status_idx)
            status = cell_status.value
            
            if status == 'OK':
                cell_status.fill = fill_ok
                cell_status.font = font_ok
            elif status == 'DIVERGÊNCIA DE VALOR':
                cell_status.fill = fill_div
                cell_status.font = font_div
                # Se tiver a coluna DIFERENCA, pinta ela de vermelho também para chamar atenção
                if coluna_diferenca_idx:
                    ws.cell(row=row, column=coluna_diferenca_idx).fill = fill_div
                    ws.cell(row=row, column=coluna_diferenca_idx).font = font_div
            elif status == 'AP NÃO ENCONTRADA NA BASE':
                cell_status.fill = fill_nao_enc
                cell_status.font = font_nao_enc
    
    # Congelar a primeira linha (cabeçalho) para facilitar a rolagem
    ws.freeze_panes = 'A2'

def processar_extracao_aps_transitarias(valores_aps: list):

    try:
        df = pd.read_excel(CAMINHO_ARQUIVO_ENTRADA)
    except FileNotFoundError:
        print(Fore.RED + f"ERRO: Arquivo de entrada não encontrado em '{CAMINHO_ARQUIVO_ENTRADA}'")
        return None, valores_aps
    if 'APS' not in df.columns:
        print(Fore.RED + "ERRO: A coluna 'APS' não foi encontrada no arquivo.")
        return None, valores_aps
    colunas_necessarias = ['NATUREZA', 'CUSTO OU DESPESA']
    df['APS'] = df['APS'].astype(str)
    valores_aps_str = [str(v).strip() for v in valores_aps]
    df_filtrado = df[df['APS'].isin(valores_aps_str)].copy()
    valores_encontrados = list(df_filtrado['APS'].unique())
    valores_nao_encontrados = [v for v in valores_aps_str if v not in valores_encontrados]
    if all(col in df_filtrado.columns for col in colunas_necessarias):
        print(Fore.MAGENTA + "Aplicando mapeamento de contas contábeis...")
        df_filtrado['conta contabil'] = df_filtrado.apply(_determinar_conta_contabil, axis=1)
        print(Fore.GREEN + "Coluna 'conta contabil' criada com sucesso.")
    return df_filtrado, valores_nao_encontrados

def salvar_dataframe_transitarias(df_principal, df_pendencias=None, df_overview=None, filename="Transitorias_Conciliadas.xlsx"):
    """
    Salva múltiplos DataFrames em um arquivo Excel com múltiplas abas.
    
    Parâmetros:
    - df_principal: DataFrame com o detalhamento dos dados conciliados
    - df_pendencias: DataFrame com as APs não encontradas
    - df_overview: DataFrame com resumo executivo
    - filename: Nome do arquivo a ser salvo
    """
    pasta_saida = str(CAMINHO_PASTA_SAIDA)
    os.makedirs(pasta_saida, exist_ok=True)
    
    caminho_saida = os.path.join(pasta_saida, filename)
    print(Fore.YELLOW + f"⏳ Gerando arquivo Excel em: {caminho_saida}..." + Style.RESET_ALL)
    
    try:
        # Usar ExcelWriter para escrever múltiplas abas
        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
            # A aba de Overview será a primeira (a que abre ao abrir o arquivo)
            if df_overview is not None:
                df_overview.to_excel(writer, sheet_name='OVERVIEW', index=False)
            
            df_principal.to_excel(writer, sheet_name='DETALHAMENTO_CONCILIADO', index=False)
            
            if df_pendencias is not None and not df_pendencias.empty:
                df_pendencias.to_excel(writer, sheet_name='APS_NAO_ENCONTRADAS', index=False)
        
        # --- Parte de Estilização (Opcional) ---
        wb = load_workbook(caminho_saida)
        
        # Aplicamos a formatação de cores na aba de detalhamento
        ws_detalhe = wb['DETALHAMENTO_CONCILIADO']
        _aplicar_estilo_transitorias(ws_detalhe)
        
        # Ajustar largura das colunas no Overview para ficar legível
        if df_overview is not None:
            ws_ov = wb['OVERVIEW']
            for col in ws_ov.columns:
                ws_ov.column_dimensions[col[0].column_letter].width = 25

        wb.save(caminho_saida)
        print(Fore.GREEN + f"✅ Relatório gerado com sucesso!" + Style.RESET_ALL)

    except PermissionError:
        print(Fore.RED + f"❌ Erro de permissão: O arquivo '{filename}' está aberto. Feche-o e tente novamente." + Style.RESET_ALL)
    except Exception as e:
        print(Fore.RED + f"❌ Erro na gravação do Excel: {e}" + Style.RESET_ALL)
# --- PASSO 2: NOVO MOTOR DE CONCILIAÇÃO FISCAL ---

def processar_conciliacao_fiscal(caminho_fiscal: str):
    """
    Lê o ficheiro do fiscal, cruza com a base de APs, valida os valores totais
    e devolve o DataFrame com as linhas de rateio classificadas e com status de conciliação.
    """
    # 1. Ler o ficheiro do fiscal
    try:
        df_fiscal = pd.read_excel(caminho_fiscal)
    except Exception as e:
        print(Fore.RED + f"Erro ao ler o ficheiro fiscal: {e}")
        return None, {}

    # Normalizar nomes das colunas (remover espaços e colocar em maiúsculas) para evitar erros
    df_fiscal.columns = df_fiscal.columns.str.strip().str.upper()

    # Validar se as colunas essenciais existem no ficheiro do fiscal
    if 'AP' not in df_fiscal.columns or 'VALOR' not in df_fiscal.columns:
        print(Fore.RED + "ERRO: O ficheiro Fiscal deve conter as colunas 'AP' e 'VALOR'.")
        print(Fore.YELLOW + f"Colunas encontradas: {list(df_fiscal.columns)}")
        return None, {}

    # 2. Ler a base de dados local
    try:
        df_base = pd.read_excel(CAMINHO_ARQUIVO_ENTRADA)
    except FileNotFoundError:
        print(Fore.RED + f"ERRO: Base de dados não encontrada em '{CAMINHO_ARQUIVO_ENTRADA}'")
        return None, {}

    if 'APS' not in df_base.columns:
        print(Fore.RED + "ERRO: A coluna 'APS' não foi encontrada na base local.")
        return None, {}

    # Identificar a coluna de valor na base local (assume-se 'VALOR')
    coluna_valor_base = 'VALOR'
    if coluna_valor_base not in df_base.columns:
        # Se não houver 'VALOR', tenta encontrar colunas parecidas (ex: 'VALOR TOTAL')
        colunas_possiveis = [c for c in df_base.columns if 'VALOR' in str(c).upper()]
        if colunas_possiveis:
            coluna_valor_base = colunas_possiveis[0]
        else:
            print(Fore.RED + "ERRO: Não foi possível identificar a coluna de valor na base local.")
            return None, {}

    # Normalizar as chaves (AP/APS) para texto (string) para garantir que o cruzamento funciona
    df_fiscal['AP'] = df_fiscal['AP'].astype(str).str.strip()
    df_base['APS'] = df_base['APS'].astype(str).str.strip()
    
    # Converter valores para numérico (substitui letras ou erros por 0)
    df_fiscal['VALOR'] = pd.to_numeric(df_fiscal['VALOR'], errors='coerce').fillna(0)
    df_base[coluna_valor_base] = pd.to_numeric(df_base[coluna_valor_base], errors='coerce').fillna(0)

    # 3. Filtrar a base local usando APENAS as APs pedidas pelo fiscal
    lista_aps_fiscal = df_fiscal['AP'].unique()
    df_base_filtrada = df_base[df_base['APS'].isin(lista_aps_fiscal)].copy()

    # 4. Agrupar (Group By) os valores das linhas de rateio da base local por AP
    df_base_agrupada = df_base_filtrada.groupby('APS', as_index=False)[coluna_valor_base].sum()
    df_base_agrupada.rename(columns={coluna_valor_base: 'VALOR_TOTAL_BASE'}, inplace=True)

    # 5. Cruzar (Merge) os totais do sistema com os totais do fiscal
    df_conciliacao = pd.merge(df_fiscal, df_base_agrupada, left_on='AP', right_on='APS', how='left')

    # 6. Validar matematicamente as diferenças
    tolerancia = 0.05 # Aceita até 5 cêntimos de diferença por arredondamento
    
    # Se VALOR_TOTAL_BASE for nulo, significa que o fiscal pediu uma AP que ainda não está na base
    df_conciliacao['STATUS_CONCILIACAO'] = 'AP NÃO ENCONTRADA NA BASE'
    
    # Isolar as APs que foram encontradas na base
    mask_encontrada = df_conciliacao['VALOR_TOTAL_BASE'].notna()
    
    # Calcular a diferença: Total da Base - Total do Fiscal
    df_conciliacao.loc[mask_encontrada, 'DIFERENCA'] = df_conciliacao.loc[mask_encontrada, 'VALOR_TOTAL_BASE'] - df_conciliacao.loc[mask_encontrada, 'VALOR']
    
    # Classificar: OK ou Divergência
    mask_ok = mask_encontrada & (df_conciliacao['DIFERENCA'].abs() <= tolerancia)
    df_conciliacao.loc[mask_ok, 'STATUS_CONCILIACAO'] = 'OK'
    
    mask_divergente = mask_encontrada & (df_conciliacao['DIFERENCA'].abs() > tolerancia)
    df_conciliacao.loc[mask_divergente, 'STATUS_CONCILIACAO'] = 'DIVERGÊNCIA DE VALOR'

    # 7. Devolver a classificação para as linhas detalhadas (Rateio)
    # Criamos dicionários para mapear os resultados de volta para as múltiplas linhas
    dict_status = dict(zip(df_conciliacao['AP'], df_conciliacao['STATUS_CONCILIACAO']))
    dict_valor_fiscal = dict(zip(df_conciliacao['AP'], df_conciliacao['VALOR']))

    print(Fore.MAGENTA + "A aplicar mapeamento de contas contabilísticas...")
    colunas_necessarias = ['NATUREZA', 'CUSTO OU DESPESA']
    if all(col in df_base_filtrada.columns for col in colunas_necessarias):
        df_base_filtrada['conta contabil'] = df_base_filtrada.apply(_determinar_conta_contabil, axis=1)

    # Injetar as colunas novas de auditoria no ficheiro final
    df_base_filtrada['VALOR_FISCAL_RAZAO'] = df_base_filtrada['APS'].map(dict_valor_fiscal)
    df_base_filtrada['STATUS_CONCILIACAO'] = df_base_filtrada['APS'].map(dict_status)

    # 8. Calcular as métricas para o resumo do terminal
    aps_encontradas = df_base_filtrada['APS'].nunique()
    aps_com_divergencia = (df_conciliacao['STATUS_CONCILIACAO'] == 'DIVERGÊNCIA DE VALOR').sum()
    aps_nao_encontradas = (df_conciliacao['STATUS_CONCILIACAO'] == 'AP NÃO ENCONTRADA NA BASE').sum()

    resumo = {
        'total_aps_fiscais': len(lista_aps_fiscal),
        'encontradas': aps_encontradas,
        'nao_encontradas': aps_nao_encontradas,
        'divergencias': aps_com_divergencia
    }

    # 9. Gerar o DataFrame de Overview (Resumo por Colunas)
    # Agrupamos por status para ver a contagem e a soma financeira
    df_overview = df_conciliacao.groupby('STATUS_CONCILIACAO').agg(
        Qtd_APs=('AP', 'count'),
        Valor_Total_Fiscal=('VALOR', 'sum'),
        Valor_Total_Sistema=('VALOR_TOTAL_BASE', 'sum')
    ).reset_index()
    
    # Adicionamos uma coluna de Diferença para facilitar a análise
    df_overview['Diferenca_Absoluta'] = df_overview['Valor_Total_Sistema'] - df_overview['Valor_Total_Fiscal']

    # 10. Separar as APs não encontradas para a aba específica de pendências
    df_nao_encontradas = df_conciliacao[df_conciliacao['STATUS_CONCILIACAO'] == 'AP NÃO ENCONTRADA NA BASE'].copy()
    df_nao_encontradas = df_nao_encontradas[['AP', 'NF', 'FORNECEDOR', 'VALOR', 'STATUS_CONCILIACAO']]

    # Retornamos agora os TRÊS DataFrames + o dicionário de resumo para o terminal
    return df_base_filtrada, df_nao_encontradas, df_overview, resumo
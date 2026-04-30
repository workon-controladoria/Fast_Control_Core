import pandas as pd
import os
from datetime import datetime
import calendar
from colorama import Fore, Style
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

MAP_EMPRESAS = {
    'INSTORE': 1968,
    'JOB': 1964,
    'CENTRO': 1970,
    'NORTE': 1969,
    'TRADE': 1965,
    'PEOPLE': 1967
}

MAP_NOMES_FILIAIS = {
    1968: 'INSTORE',
    1964: 'JOB',
    1970: 'CENTRO',
    1969: 'NORTE',
    1965: 'TRADE',
    1967: 'PEOPLE'
}

MAP_CONTAS = {
    'VALOR BRUTO': ('12651', '30051', 'SERVICOS A FATURAR'),
    'PIS': ('30061', '20432', 'PIS'),
    'COFINS': ('30062', '20433', 'COFINS'),
    'VALOR ISS': ('30063', '20431', 'ISS')
}

def obter_datas_provisao(ano: int, mes: int):
    """Calcula o último dia do mês (Provisão) e o 1º dia do mês seguinte (Reversão)."""
    ultimo_dia = calendar.monthrange(ano, mes)[1]

    data_provisao = datetime(ano, mes, ultimo_dia).strftime('%d/%m/%Y')

    if mes == 12:
        data_reversao = datetime(ano + 1, 1, 1).strftime('%d/%m/%Y')
    else:
        data_reversao = datetime(ano, mes + 1, 1).strftime('%d/%m/%Y')
        
    return data_provisao, data_reversao

def processar_planilha_provisoes(caminho_arquivo: str, ano: int, mes: int):
    """Lê a planilha horizontal, aplica o De-Para e gera as linhas verticais do ERP."""
    print(Fore.YELLOW + f" Lendo a base de provisões: {caminho_arquivo}...")
    
    try:
        df_input = pd.read_excel(caminho_arquivo, header=None)
        linha_cabecalho = df_input[df_input.apply(lambda x: x.astype(str).str.contains('CLIENTE', case=False).any(), axis=1)].index[0]
        df_input.columns = df_input.iloc[linha_cabecalho].str.strip().str.upper()
        df_input = df_input.iloc[linha_cabecalho + 1:].copy()
        df_input.dropna(subset=['CLIENTE'], inplace=True)
        
    except Exception as e:
        print(Fore.RED + f"Erro ao ler a planilha: {e}")
        return None

    data_provisao, data_reversao = obter_datas_provisao(ano, mes)
    linhas_saida = []

    for _, row in df_input.iterrows():
        cliente = str(row.get('CLIENTE', '')).strip()
        cc = str(row.get('C.C.', '')).strip()
        empresa_nome = str(row.get('EMPRESA', '')).strip().upper()
        
        codigo_filial = MAP_EMPRESAS.get(empresa_nome, '')
        if not codigo_filial:
            continue

        for coluna_origem, (conta_deb, conta_cred, nome_hist) in MAP_CONTAS.items():
            valor = row.get(coluna_origem, 0)
            
            try:
                valor = float(valor)
            except (ValueError, TypeError):
                valor = 0
                
            if pd.isna(valor) or valor == 0:
                continue

            hist_prov = f"{nome_hist} REF. {cliente} CC {cc}"
            cc_deb_prov = ""
            cc_cred_prov = ""

            if conta_deb == '12651':
                cc_cred_prov = cc
            else:
                cc_deb_prov = cc

            linhas_saida.append({
                'Data': data_provisao,
                'Cód. Conta Debito': conta_deb,
                'Cód. Conta Credito': conta_cred,
                'Valor': round(valor, 2),
                'Cód. Histórico': 0,
                'Complemento Histórico': hist_prov,
                'Inicia Lote': 1,
                'Código Matriz/Filial': codigo_filial,
                'Centro de Custo Débito': cc_deb_prov,
                'Centro de Custo Crédito': cc_cred_prov
            })

            hist_rev = f"REVERSÃO - {nome_hist} REF. {cliente} CC {cc}"

            cc_deb_rev = cc_cred_prov
            cc_cred_rev = cc_deb_prov

            linhas_saida.append({
                'Data': data_reversao,
                'Cód. Conta Debito': conta_cred,  # Conta Invertida
                'Cód. Conta Credito': conta_deb,  # Conta Invertida
                'Valor': round(valor, 2),
                'Cód. Histórico': 0,
                'Complemento Histórico': hist_rev,
                'Inicia Lote': 1,
                'Código Matriz/Filial': codigo_filial,
                'Centro de Custo Débito': cc_deb_rev,   # C.C. Invertido
                'Centro de Custo Crédito': cc_cred_rev  # C.C. Invertido
            })

    df_final = pd.DataFrame(linhas_saida)
    df_final.sort_values(by=['Data', 'Código Matriz/Filial', 'Complemento Histórico'], inplace=True)
    
    return df_final

def salvar_layout_erp_por_empresa(df: pd.DataFrame, filename: str = "Provisoes_Layout_ERP.xlsx"):
    """
    Pega o DataFrame vertical e salva em um único arquivo, 
    separando cada empresa em uma aba (sheet) diferente.
    """
    from ..config.settings import OUTPUT_DIR
    
    caminho_saida = OUTPUT_DIR / filename
    
    try:

        with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:

            empresas_presentes = df['Código Matriz/Filial'].unique()
            
            for cod_filial in empresas_presentes:
                nome_aba = MAP_NOMES_FILIAIS.get(cod_filial, f"Filial_{cod_filial}")

                df_empresa = df[df['Código Matriz/Filial'] == cod_filial].copy()

                df_empresa.to_excel(writer, sheet_name=nome_aba, index=False)

        wb = load_workbook(caminho_saida)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            ws.freeze_panes = 'A2'
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2

        wb.save(caminho_saida)
        print(Fore.GREEN + f"✅ Layout ERP gerado com sucesso! Abas criadas: {', '.join(wb.sheetnames)}")
        print(Fore.WHITE + f"📍 Local: {caminho_saida}")

    except Exception as e:
        print(Fore.RED + f"❌ Erro ao salvar o arquivo por abas: {e}")